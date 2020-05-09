import random
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

#Klasa obiektu Cube które nic nie robi póki co
#class Square:
#    def __init__(self):
#        print('test')
Square_array = []
def Square(i):
    for i in range(1,i):
        for j in range(1,i):
            Square_array.append()


#Funkcja problem która wstawia w losowe miejsca niewiadome.
def problem(i,points_v2_normal):
    while i > 0:
        x = random.randint(0,len(points_v2_normal)-1)
        y = random.randint(0,len(points_v2_normal)-1)
        if(points_v2_normal[x][y] == 'x'):
            continue
        else:
            points_v2_normal[x][y] = 'x'
            i = i - 1
    print('Generated problem: ', *points_v2_normal,sep='\n',end='\n\n')

#Funkcja rozwiązania która próbuje z problemu osiągnąć stan na wejściu
def solution_v2(points_v2_normal):

    for points in points_v2_normal:
        for point in points:
            #Jeżeli element to string to sprawdzam jaki jest rozmiar listy i pętlą for sprawdzam ile dana liczba razy wystąpiła 
            # i jeśli jedna liczba nie wystąpiła to jest ona tą zmienną.
            if(type(point) is str):
                elements = len(points)
                for element in range(1,int(elements)+1):
                    if(points.count(element) < 1):
                        try:
                            #print(str(element) + ' is a x')
                            if((points_v2_normal[(points_v2.index(points)+1)%3][points.index(point)] == element) or (points_v2_normal[(points_v2.index(points)+2)%3][points.index(point)] == element)):
                                if(((element + 1)%4) == 0):
                                    points_v2_normal[points_v2.index(points)][points.index(point)] = 1
                                else:
                                    if((points_v2_normal[(points_v2.index(points)+1)%3][points.index(point)] == element + 1) or (points_v2_normal[(points_v2.index(points)+2)%3][points.index(point)] == element + 1)):
                                        if(((element + 1)%4) == 0):
                                            points_v2_normal[points_v2.index(points)][points.index(point)] = 1
                                        elif(element + 2 == 4):
                                            points_v2_normal[points_v2.index(points)][points.index(point)] = 1
                                        else:
                                            points_v2_normal[points_v2.index(points)][points.index(point)] = (element + 2)
                                    else:
                                        points_v2_normal[points_v2.index(points)][points.index(point)] = (element + 1)
                            else:
                                points_v2_normal[points_v2.index(points)][points.index(point)] = element
                            #possible_elements_normal.append(element)
                            #new_points_normal.append(element)
                        except ValueError:
                            print('VALUE ERROR, RETRYING...')
                            continue


#    for points in points_transposed:
#        for point in points:
#            if(type(point) is str):
#                elements_transposed = len(points)
#                for element_transposed in range(1,int(elements_transposed)+1):
#                    if(points.count(element_transposed) < 1):
#                        print(str(element_transposed) + ' is a x')
#                        points_v2_transposed[points_transposed.index(points)][points.index(point)] = element_transposed
#                        #possible_elements_transposed.append(element_transposed)
#                        #new_points_transposed.append(element_transposed)

    #Dla normal
    print('Normal: ', *points_v2_normal, sep='\n', end='\n\n')
    #for points in points_v2_normal_reversed:
    #    for point in points:
    #        if(points.count(point) > 1):
    #            print('Error', point)
    #            Status_normal_boolean = True
                #print(points_v2_normal_reversed.index(points))
                #point = 'x'
                #print(points_v2_normal_reversed)
                #solution_v2()
    #print(*points_v2_normal_reversed, sep='\n', end='\n\n')
    points_v2_transposed = list(map(list, zip(points_v2_normal[0],points_v2_normal[1],points_v2_normal[2])))
    #Dla transposed
    #print('Transposed: ', *points_v2_transposed, sep='\n', end='\n\n')    
    #for points in points_v2_transposed_reversed:
    #    for point in points:
    #        if(points.count(point) > 1):
    #            print('Error', point)
    #            Status_transposed_boolean = True
                #point = 'x'
                #print(points_v2_transposed_reversed)
                #solution_v2
    #print(*points_v2_transposed_reversed, sep='\n', end='\n\n')


#Funkcja która nic nie robi póki co
def solution_t(points_v2_normal,problem_amount):
    i = 0
    j = 0
    for points in points_v2_normal:
        for point in points:
            elements = len(points)
            if(points.count(point) == 1):
                i = i + 1
            else:
                i = i + (1/2)
    points_v2_normal_reversed = list(map(list, zip(points_v2_normal[0],points_v2_normal[1],points_v2_normal[2])))
    for points in points_v2_normal_reversed:
        for point in points:
            elements = len(points)
            if(points.count(point) == 1):
                j = j + 1
            else:
                j = j + (1/2)

    score = i + j
    score = round((i + j)/(amount_of_elements*2),2)*100
    print(score,'%',sep='')
    return score

def load_from_file():
    text_file_source = open('Source.txt', 'r')
    for line in text_file_source:
        stripped_line = line.strip()
        line_list = stripped_line.split()
        points_from_txt.append(line_list)
    print(points_from_txt)


#Listy na kompletne listy z liczbami zamiast zmiennej
new_points_normal = []
new_points_transposed = []

#Listy z możliwymi liczbami które można podstawić za zmienną
possible_elements_normal = []
possible_elements_transposed = []

#Tworzenie setów do porównywania czy jeden zbiór jest podzbiorem drugiego
set_normal = set(possible_elements_normal)
set_transposed = set(possible_elements_transposed)

#Listy z danymi wejściowymi, danymi po transpozycji oraz liczbą elementów w liście
#points_v2 = [[2, 1, 3],[3, 2, 1],[1, 3, 2]]
#points_v2 = [[1,2,3,4,5,6,7,8],[2,3,4,5,6,7,8,1],[3,4,5,6,7,8,1,2],[4,5,6,7,8,1,2,3],[5,6,7,8,1,2,3,4],[6,7,8,1,2,3,4,5],[7,8,1,2,3,4,5,6],[8,1,2,3,4,5,6,7]]

#points_transposed = list(map(list, zip(points_v2[0],points_v2[1],points_v2[2],points_v2[3],points_v2[4],points_v2[5],points_v2[6],points_v2[7])))


#Kopia listy do ocenienia jakości oraz kopia listy z niewiadomymi oraz kopia listy po transpozycji

#Wywołanie funkcji
problem_amount = 0
solution_amount = 0
points_v2 = [[1,2,3],[2,3,1],[3,1,2]]
points_from_txt = []
points_for_excel = []
columns_table = ['G','H','I']
rows_table = [5,6,7]

#points_transposed = list(map(list, zip(points_v2[0],points_v2[1],points_v2[2])))
amount_of_elements = len(points_v2)*len(points_v2)
#points_v2_normal = points_v2
#points_v2_transposed = points_transposed

#solution_v2()
#points_v2_normal_reversed = list(map(list, zip(points_v2_normal[0],points_v2_normal[1],points_v2_normal[2])))
#solution_t()
#after = datetime.now()

trial_amount = int(input('How many test cases?: '))
problem_amount = int(input('How many unknown values do you want?: '))
source = int(input('Select data source:\n1-From txt file\n2-Default source data\n'))

def trial(trial_amount, problem_amount):
    interval_sum = 0
    i = 1
    if(source == 1):
        load_from_file()
    text_file = open('Result.txt', 'w')

    wb = Workbook()
    ws1 = wb.active
    
    ws1.title = 'Results'
    ws1['A1'] = 'ID'
    ws1['B1'] = 'SCORE'
    ws1['C1'] = 'TIME'
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 20
    ws1.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws1.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws1.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for each in range(0, trial_amount):
        ws = wb.create_sheet("Result ", each)
        before = datetime.now()
        points_v2_normal = points_v2
        problem(problem_amount,points_v2_normal)
        for row in points_v2_normal:
            text_file.write(str(row) + '\n')
        text_file.write('\n')
        solution_v2(points_v2_normal)
        points_v2_normal_reversed = list(map(list, zip(points_v2_normal[0],points_v2_normal[1],points_v2_normal[2])))
        score = solution_t(points_v2_normal,problem_amount)
        after = datetime.now()
        interval = after - before
        if(type(interval_sum) == int):
            interval_sum = interval
        else:
            interval_sum = interval_sum + interval
        text_file.write('Latin square: \n')
        for row in points_v2_normal:
            text_file.write(str(row) + '\n')
        text_file.write('\n')
        ws1['A' + str(i + 1)] = i
        ws1['B' + str(i + 1)] = str(score) + '%'
        ws1['C' + str(i + 1)] = str(interval)
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20

        ws.row_dimensions[5].height = 80
        ws.row_dimensions[6].height = 80
        ws.row_dimensions[7].height = 80
        rows = range(5,8)
        columns = range(7,10)
        for row in rows:
            for col in columns:
                ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for points in points_v2_normal:
            for point in points:
                points_for_excel.append(point)
        pointer = 0
        for each_c in columns_table:
            for each_r in rows_table:
                ws[str(each_c) + str(each_r)] = points_for_excel[pointer]
                pointer = pointer + 1
        ws.conditional_formatting.add('G5:I7',ColorScaleRule(start_type='percentile', start_value=10, start_color='AA0000',mid_type='percentile', mid_value=50, mid_color='F7A311',end_type='percentile', end_value=90, end_color='00AA00'))
        print(interval)
        print(i)
        i = i + 1
        print(points_v2)
    text_file.write('Total time: ' + str(interval_sum) + '\nScore: ' + str(score) + '%\n\n')
    text_file.close()
    print(interval_sum)
    wb.save('Results.xlsx')

trial(trial_amount,problem_amount)
#175000 to ok 10 minut